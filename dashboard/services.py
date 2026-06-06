import openpyxl
from datetime import datetime
from django.utils import timezone
import re
from .models import NovedadEstadistica, AmpliacionEstadistica

def limpiar_int(val):
    if val is None or str(val).strip() == "" or str(val).lower() == 'null': return 0
    try:
        # Maneja casos donde el número viene como "1.0" o "1,0"
        return int(float(str(val).replace(',', '.')))
    except: return 0

def limpiar_fecha_excel(cell):
    if cell is None or str(cell).lower() == 'null': return None
    if isinstance(cell, datetime):
        if timezone.is_naive(cell):
            return timezone.make_aware(cell, timezone.get_current_timezone())
        return cell
    # Si viene como string (ej: "01/03/2025 10:00")
    try:
        txt = str(cell).strip()
        if '/' in txt:
            dt = datetime.strptime(txt, '%d/%m/%Y %H:%M')
            return timezone.make_aware(dt, timezone.get_current_timezone())
    except:
        return None
    return None

def obtener_alias(nombre_largo):
    if not nombre_largo: return "OTRA"
    n = str(nombre_largo).upper().replace('–', '-').strip()
    MAPEO = {
        "AYUDANTÍA": "AYUDANTÍA", "DELITOS ESPECIALES": "DIDE", 
        "ANÁLISIS Y REG": "DARI", "GUARDIA": "GUARDIA", 
        "FLOTANTE": "DPF", "FINANCIEROS": "DIDF", 
        "TRÁF AUTOMOTOR": "DATA", "LAVADO DE ACTIVOS": "SLA", 
        "I24/7": "I24/7", "CAPTURAS": "DCI", 
        "ASUNTOS INTERNACIONALES": "DAI", "PERSONAS AUSENTES": "DRBPA"
    }
    for clave, alias in MAPEO.items():
        if clave in n: return alias
    return "OTRA"

def procesar_excel_eventos(archivo):
    wb = openpyxl.load_workbook(archivo, data_only=True)
    sheet = wb.active
    
    # 1. Mapeo de cabeceras (Limpiamos puntos, tildes y espacios)
    def normalizar(t): return str(t).lower().replace('.', '').replace('í', 'i').strip() if t else ""
    
    headers = {normalizar(cell.value): i for i, cell in enumerate(sheet[1]) if cell.value}
    
    # Verificación de columnas críticas
    required = ['nro', 'ingreso', 'dependencia actual']
    for req in required:
        if req not in headers:
            raise ValueError(f"No se encontró la columna '{req}' en el Excel. Columnas detectadas: {list(headers.keys())}")

    nuevos, actualizados = 0, 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nro_nov = row[headers['nro']]
        if not nro_nov: continue

        dep_larga = row[headers['dependencia actual']] or ''

        # Usamos .get() con el índice para mayor seguridad
        obj, created = NovedadEstadistica.objects.update_or_create(
            nro=nro_nov,
            defaults={
                'nunc': row[headers.get('noticiaunicacriminal')] if 'noticiaunicacriminal' in headers else None,
                'fecha_ingreso': limpiar_fecha_excel(row[headers['ingreso']]),
                'dependencia_original': dep_larga,
                'dependencia_alias': obtener_alias(dep_larga),
                'titulo': row[headers.get('titulo')] or 'SIN TÍTULO' if 'titulo' in headers else 'SIN TÍTULO',
                'aclarada': str(row[headers.get('aclarada', -1)]).lower() == 'si' if 'aclarada' in headers else False,
                'allanamientos_pos': limpiar_int(row[headers.get('total allanamientos positivos')]) if 'total allanamientos positivos' in headers else 0,
                'allanamientos_neg': limpiar_int(row[headers.get('total allanamientos negativos')]) if 'total allanamientos negativos' in headers else 0,
            }
        )
        if created: nuevos += 1
        else: actualizados += 1
    return nuevos, actualizados

def procesar_excel_ampliaciones_masivo(lista_archivos):
    total = 0
    patron_fecha = r'(\d{4}-\d{2})'

    for archivo in lista_archivos:
        match = re.search(patron_fecha, archivo.name)
        if not match: continue
        periodo = match.group(1)
        
        wb = openpyxl.load_workbook(archivo, data_only=True)
        sheet = wb.active
        
        def normalizar(t): return str(t).lower().replace('.', '').replace('í', 'i').strip() if t else ""
        headers = {normalizar(cell.value): i for i, cell in enumerate(sheet[1]) if cell.value}

        if 'nro' not in headers or 'denuncia' not in headers:
            continue # Salta archivos que no tengan las columnas básicas

        for row in sheet.iter_rows(min_row=2, values_only=True):
            nro_amp = row[headers['nro']]
            if not nro_amp: continue

            dep_larga = row[headers.get('dependencia ampliacion')] or ''

            AmpliacionEstadistica.objects.update_or_create(
                nro=nro_amp,
                defaults={
                    'denuncia_madre': row[headers.get('denuncia')],
                    'fecha_denuncia': limpiar_fecha_excel(row[headers.get('ingreso denuncia')]),
                    'titulo': row[headers.get('titulo')] or 'SIN TÍTULO' if 'titulo' in headers else 'SIN TÍTULO',
                    'dependencia_alias': obtener_alias(dep_larga),
                    'tipo': row[headers.get('tipo')] or 'AMP' if 'tipo' in headers else 'AMP',
                    'periodo': periodo,
                    'allanamientos_pos': limpiar_int(row[headers.get('total allanamientos positivos')]) if 'total allanamientos positivos' in headers else 0,
                    'allanamientos_neg': limpiar_int(row[headers.get('total allanamientos negativos')]) if 'total allanamientos negativos' in headers else 0,
                }
            )
            total += 1
    return total